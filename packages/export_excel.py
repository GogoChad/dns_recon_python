"""Export DNS results to Excel spreadsheets."""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


def export_excel(data, output_file='dns_map.xlsx'):
    if openpyxl is None:
        raise ImportError("openpyxl package not installed. Install with: pip install openpyxl")
    
    # Create new workbook and remove default sheet
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create summary sheet with scan metadata
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "DNS Mapping Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A3'] = "Domain:"
    ws_summary['B3'] = data.get('domain', 'N/A')
    ws_summary['A4'] = "Scan Date:"
    ws_summary['B4'] = data.get('scan_date', 'N/A')
    ws_summary['A5'] = "Total Results:"
    ws_summary['B5'] = data.get('total_results', 0)
    
    # Define consistent styling for all sheet headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Create sheets for each DNS strategy
    results = data.get('results', {})
    for strategy_name, strategy_results in results.items():
        # Create new sheet for this strategy (uppercase for clarity)
        ws = wb.create_sheet(strategy_name.upper())
        
        # Add column headers
        ws['A1'] = "Type"
        ws['B1'] = "Value"
        ws['C1'] = "Details"
        
        # Apply header styling
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
        
        # Populate data rows
        if isinstance(strategy_results, list):
            for idx, result in enumerate(strategy_results, start=2):
                if isinstance(result, dict):
                    # Extract structured data from dict
                    ws[f'A{idx}'] = result.get('type', strategy_name)
                    ws[f'B{idx}'] = result.get('value', str(result))
                    ws[f'C{idx}'] = result.get('details', '')
                else:
                    # Handle simple string results
                    ws[f'A{idx}'] = strategy_name
                    ws[f'B{idx}'] = str(result)
        
        # Auto-adjust column widths for readability
        for column in ['A', 'B', 'C']:
            ws.column_dimensions[column].width = 30
    
    # Save workbook to file
    wb.save(output_file)
    return output_file
